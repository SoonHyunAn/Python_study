from openpyxl import Workbook
wb = Workbook()
ws = wb.active

# 병합하기
ws.merge_cells("B2:D2") # B2 부터 D2 까지 합치겠음
ws["B2"].value = "Merged Cell"

wb.save("sample_merge.xlsx")


from openpyxl import load_workbook
wb = load_workbook("sample_merge.xlsx")
ws = wb.active

# B2:D2 병합되어 있던 셀을 해제
ws.unmerge_cells("B2:D2")
wb.save("sample_unmerge.xlsx")
