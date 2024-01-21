from openpyxl import Workbook

wb = Workbook()
ws = wb.create_sheet() # 새로운 시트 (이름은 기본)
ws.title = "내 시트" #시트 이름
ws.sheet_properties.tabColor = "ff66ff" # 시트 색 지정 (RGB)

ws1=wb.create_sheet("네 시트")
ws2=wb.create_sheet("우리 시트",2) # 인덱스가 2번인곳에 생성

new_ws=wb["우리 시트"]

print(wb.sheetnames) # 모든 시트 이름 리스트

# sheet 복사
new_ws["A1"] = "test" # new_ws의 시트 이름이 "우리 시트"로 지정해놓았기 때문에 "우리 시트"와 "복사된 시트"의 A1정보가 동일하다.
target=wb.copy_worksheet(new_ws)
target.title = "복사된 시트"

wb.save("Sample.xlsx")
wb.close()