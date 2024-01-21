from openpyxl import load_workbook # 파일 불러오기
wb = load_workbook("sample.xlsx") # sample.xlsx 파일에서 wb 을 불러옴
ws = wb.active # 활성화된 Sheet

# 찾기, 수정
for row in ws.iter_rows(min_row=2):
    # 번호, 영어, 수학
    if int(row[1].value) > 80:
        print(row[0].value, "번 학생은 영어 천재")

for row in ws.iter_rows(max_row=1):
    for cell in row:
        if cell.value == "영어":
            cell.value = "컴퓨터"

wb.save("sample_modified.xlsx")


# 행과 열 추가
ws.insert_rows(8) # 8번째 줄이 비워짐 (새로운 빈 행이 추가)
ws.insert_rows(8, 5) # 8번째 줄 위치에 5줄을 추가
wb.save("sample_insert_rows.xlsx")

ws.insert_cols(2) # B번째 열이 비워짐 (새로운 빈 열이 추가)
ws.insert_cols(2, 3) # B번째 열로부터 3열 추가
wb.save("sample_insert_cols.xlsx")


# 삭제
ws.delete_rows(8) # 8 번째 줄에 있는 7 번 학생 데이터 삭제
ws.delete_rows(8, 3) # 8번째 줄부터 총 3줄 삭제
wb.save("sample_delete_row.xlsx")

ws.delete_cols(2) # 2번째 열 (B) 삭제
ws.delete_cols(2, 2) # 2번재 열로부터 총 2개 열 삭제
wb.save("sample_delete_col.xlsx")


#이동
# 번호 영어 수학
# 번호 (국어) 영어 수학 꼴을 위해 값 이동
ws.move_range("B1:C11", rows=0, cols=1) # 원하는 값의 범위, Row이동 수, Columb 이동 수
ws["B1"].value = "국어" # B1 셀에 '국어' 입력

# 번호 영어 수학
ws.move_range("C1:C11", rows=5, cols=-1)

wb.save("sample_move.xlsx")