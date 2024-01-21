import datetime
from openpyxl import Workbook
wb = Workbook()
ws = wb.active

ws["A1"] = datetime.datetime.today() # 오늘 날짜 정보
ws["A2"] = "=SUM(1, 2, 3)" # 1 + 2 + 3 = 6 (합계)
ws["A3"] = "=AVERAGE(1, 2, 3)" # 2 (평균)

ws["A4"] = 10
ws["A5"] = 20
ws["A6"] = "=SUM(A4:A5)" # 30

#이외에도 다양한 함수 사용 가능

wb.save("sample_formula.xlsx")

from openpyxl import load_workbook
# wb = load_workbook("sample_formula.xlsx")
# ws = wb.active

# # 수식 그대로 가져오고 있음
# for row in ws.values:
#     for cell in row:
#         print(cell)

wb = load_workbook("sample_formula.xlsx", data_only=True)
ws = wb.active

# 수식이 아닌 실제 데이터를 가지고 옴
# evaluate 되지 않은 상태의 데이터는 None 이라고 표시
# 엑셀 파일을 열어 계산된 값을 확인하고 저장하면 함수 적용 값이 출력
for row in ws.values:
    for cell in row:
        print(cell)